# reports/xlsx.py
#
# Module for xlsx report generation functions for spdxLicenseManager.
#
# Copyright (C) The Linux Foundation
# SPDX-License-Identifier: Apache-2.0
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import os
from collections import OrderedDict
import openpyxl

from .common import ReportFileError, ReportNotReadyError
from ..projectdb import ProjectDBQueryError, License

class XlsxReporter:

  def __init__(self, db, config={}):
    super(XlsxReporter, self).__init__()
    self._reset()
    self.db = db
    # copy over config entries into new dict
    for key, value in config.items():
      self.kwConfig[key] = value

  ##### Main xlsx reporting functions
  ##### External usage shouldn't require calling anything except these

  def setResults(self, results):
    self.wb = openpyxl.Workbook()
    self.results = results
    # look for "No license found" category, and if one exists, annotate it
    for cat in self.results.values():
      if cat.name == "No license found":
        nextLicID = self._getResultsMaxLicenseID(self.results) + 1
        self._annotateNoLicenseFound(catNoLicense=cat, nextLicID=nextLicID)

  def generate(self):
    if type(self.results) != OrderedDict:
      raise ReportNotReadyError("Cannot call generate() before analysis results are set")

    strip_licenseref = False
    if self._getFinalConfigValue("report-strip-licenseref") == 'yes':
      strip_licenseref = True

    if self._getFinalConfigValue("report-include-summary") == 'yes':
      self._generateSummarySheet(self.wb, self.results, strip_licenseref=strip_licenseref)

    self._generateCategorySheets(self.wb, self.results)
    self._generateFileListings(self.wb, self.results, strip_licenseref=strip_licenseref)
    self.reportGenerated = True

  def save(self, path, replace=False):
    self._saveCheck(path=path, replace=replace)
    try:
      self.wb.save(path)
    except PermissionError:
      raise ReportFileError(f"Permission denied to save to {path}")

  ##### Helper functions for xlsx reporting

  def _generateCategorySheets(self, wb, results):
    # create font styles
    fontBold = openpyxl.styles.Font(size=16, bold=True)
    fontNormal = openpyxl.styles.Font(size=14)
    alignNormal = openpyxl.styles.Alignment(wrap_text=True)

    # if first sheet isn't a summary sheet, then the existing first sheet
    # should be renamed when we come to the first category with files
    first_sheet = (wb.sheetnames[0] != 'License summary')

    for cat in results.values():
      # skip category if it has no files
      if not cat.hasFiles:
        continue

      # change sheet name or add new sheet, as appropriate
      if first_sheet:
        ws = wb.active
        ws.title = cat.name
        first_sheet = False
      else:
        ws = wb.create_sheet(cat.name)

      # and set column dimensions
      ws.column_dimensions['A'].width = 100
      ws.column_dimensions['B'].width = 60

      # and fill in sheet headers
      ws['A1'] = "File"
      ws['A1'].font = fontBold
      ws['B1'] = "License"
      ws['B1'].font = fontBold

  def _generateFileListings(self, wb, results, strip_licenseref=False):
    # create font styles
    fontBold = openpyxl.styles.Font(size=16, bold=True)
    fontNormal = openpyxl.styles.Font(size=14)
    alignNormal = openpyxl.styles.Alignment(wrap_text=True)

    for cat in results.values():
      if not cat.hasFiles:
        continue
      try:
        ws = wb[cat.name]
      except KeyError:
        raise ReportNotReadyError(f"Sheet not found for category {cat.name}; has _generateCategorySheets() been called?")
      row = 2
      for lic in cat.licensesSorted.values():
        if not lic.hasFiles:
          continue
        for file in lic.filesSorted.values():
          ws[f'A{row}'] = file.path
          ws[f'A{row}'].font = fontNormal
          ws[f'A{row}'].alignment = alignNormal
          ws[f'B{row}'] = self._getFinalLicenseName(lic.name, strip_licenseref)
          ws[f'B{row}'].font = fontNormal
          ws[f'B{row}'].alignment = alignNormal
          row += 1

  def _generateSummarySheet(self, wb, results, strip_licenseref=False):
    # use the first (existing) sheet as the summary sheet
    ws = wb.active
    ws.title = "License summary"

    # adjust column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 10

    # create font styles
    fontBold = openpyxl.styles.Font(size=16, bold=True)
    fontNormal = openpyxl.styles.Font(size=14)
    alignNormal = openpyxl.styles.Alignment(wrap_text=True)

    # create headers
    ws['A1'] = "License"
    ws['A1'].font = fontBold
    ws['C1'] = "# of files"
    ws['C1'].font = fontBold

    # create category and license rows
    total = 0
    row = 3
    for cat in results.values():
      if not cat.hasFiles:
        continue
      ws[f'A{row}'] = f'{cat.name}:'
      ws[f'A{row}'].font = fontBold
      row += 1
      for lic in cat.licensesSorted.values():
        if not lic.hasFiles:
          continue
        numfiles = len(lic.filesSorted)
        ws[f'B{row}'] = self._getFinalLicenseName(lic.name, strip_licenseref)
        ws[f'B{row}'].font = fontNormal
        ws[f'B{row}'].alignment = alignNormal
        ws[f'C{row}'] = numfiles
        ws[f'C{row}'].font = fontNormal
        total += numfiles
        row += 1

    # create total row
    row += 1
    ws[f'A{row}'] = "TOTAL"
    ws[f'A{row}'].font = fontBold
    ws[f'C{row}'] = total
    ws[f'C{row}'].font = fontBold

  def _annotateNoLicenseFound(self, catNoLicense, nextLicID):
    # figure out which ones we're checking
    licExt = None
    checkExt = (self._getFinalConfigValue("analyze-extensions") == "yes")
    licEmpty = None
    checkEmpty = (self._getFinalConfigValue("analyze-emptyfile") == "yes")
    licThird = None
    checkThird = (self._getFinalConfigValue("analyze-thirdparty") == "yes")

    # for now, only looking at the first license in the category
    lic = list(catNoLicense.licensesSorted.values())[0]
    for file in lic.filesSorted.values():
      labeled = False
      if checkThird and not labeled:
        if file.findings.get("thirdparty", "N/A") == "yes":
          # create new "license" if this is the first file we've seen for it
          if licThird is None:
            licThird = self._createTempLicense(
              catID=catNoLicense._id,
              nextLicID=nextLicID,
              name="No license found - third party directory"
            )
            nextLicID += 1
            catNoLicense.licensesSorted[licThird._id] = licThird
          # add file to this "license"
          licThird.filesSorted[file._id] = file
          labeled = True
      if checkEmpty and not labeled:
        if file.findings.get("emptyfile", "N/A") == "yes":
          # create new "license" if this is the first file we've seen for it
          if licEmpty is None:
            licEmpty = self._createTempLicense(
              catID=catNoLicense._id,
              nextLicID=nextLicID,
              name="No license found - empty file"
            )
            nextLicID += 1
            catNoLicense.licensesSorted[licEmpty._id] = licEmpty
          # add file to this "license"
          licEmpty.filesSorted[file._id] = file
          labeled = True
      if checkExt and not labeled:
        if file.findings.get("extension", "N/A") == "yes":
          # create new "license" if this is the first file we've seen for it
          if licExt is None:
            licExt = self._createTempLicense(
              catID=catNoLicense._id,
              nextLicID=nextLicID,
              name="No license found - excluded file extension"
            )
            nextLicID += 1
            catNoLicense.licensesSorted[licExt._id] = licExt
          # add file to this "license"
          licExt.filesSorted[file._id] = file
          labeled = True

    # wait until we're done, so we can go back and remove them now
    # (can't mutate the filesSorted OrderedDict while iterating)
    if licThird is not None:
      for file in licThird.filesSorted.values():
        # remove from the original license file list
        del lic.filesSorted[file._id]
    if licEmpty is not None:
      for file in licEmpty.filesSorted.values():
        # remove from the original license file list
        del lic.filesSorted[file._id]
    if licExt is not None:
      for file in licExt.filesSorted.values():
        # remove from the original license file list
        del lic.filesSorted[file._id]

  def _saveCheck(self, path, replace=False):
    if type(self.results) != OrderedDict:
      raise ReportNotReadyError("Cannot call save() before analysis results are set")
    if not self.reportGenerated:
      raise ReportNotReadyError("Cannot call save() before report is generated")

    # check whether requested file already exists
    if os.path.exists(path) and not replace:
      raise ReportFileError(f"File already exists at {path}")

    # check whether we have write permission for this path
    if not os.access(path=os.path.dirname(path), mode=os.W_OK):
      raise ReportFileError(f"Permission denied to save to {path}")

  ##### Other helper functions

  def _reset(self):
    self.wb = None
    self.results = None
    self.reportGenerated = False
    self.kwConfig = {}

  def _getFinalConfigValue(self, key):
    kwValue = self.kwConfig.get(key, None)
    if kwValue is not None:
      return str(kwValue).lower()
    try:
      value = self.db.getConfigValue(key)
      return str(value).lower()
    except ProjectDBQueryError:
      return ""

  def _getResultsMaxLicenseID(self, results):
    maxLicID = 0
    for cat in results.values():
      for lic in cat.licensesSorted.values():
        if lic._id > maxLicID:
          maxLicID = lic._id
    return maxLicID

  def _createTempLicense(self, catID, nextLicID, name):
    lic = License()
    lic._id = nextLicID
    lic.name = name
    lic.category_id = catID
    lic.hasFiles = True
    lic.filesSorted = OrderedDict()
    return lic

  def _getFinalLicenseName(self, licName, stripLicenseRef):
    if stripLicenseRef:
      return licName.replace("LicenseRef-", "")
    else:
      return licName
