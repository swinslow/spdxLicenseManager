# tests/ft_reportxlsx.py
#
# Functional tests for spdxLicenseManager: producing Xlsx reports from
# previously-imported scans.
#
# Copyright (C) The Linux Foundation
# SPDX-License-Identifier: Apache-2.0
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import unittest
import click
from click.testing import CliRunner

from testfixtures import TempDirectory
import os
from openpyxl import load_workbook

from slm import slm

from helper_sandbox import (setUpSandbox, runSandboxCommands, tearDownSandbox,
  runcmd, printResultDebug)
from helper_check import checkForFileExists

# paths to various SPDX test files
PATH_SIMPLE_SPDX = "tests/testfiles/simple.spdx"
PATH_SIMPLE_EXTENSION_SPDX = "tests/testfiles/simpleExtension.spdx"
PATH_SIMPLE_EMPTYFILE_SPDX = "tests/testfiles/simpleEmptyFile.spdx"
PATH_SIMPLE_THIRDPARTY_SPDX = "tests/testfiles/simpleThirdParty.spdx"
PATH_SIMPLE_LICENSEREF_SPDX = "tests/testfiles/simpleLicenseRef.spdx"
PATH_SPDX_SUMMARIZER = "tests/testfiles/spdxSummarizer-test1.spdx"

class XlsxReportFuncTestSuite(unittest.TestCase):
  """spdxLicenseManager Xlsx reporting FT suite."""

  def setUp(self):
    self.runner = CliRunner()
    setUpSandbox(self, slm.cli)
    runSandboxCommands(self, slm.cli)

    # set up temp directory for outputting reports
    self.reportDir = TempDirectory()

  def tearDown(self):
    self.reportDir.cleanup()
    self.reportDir = None
    tearDownSandbox(self)

  ##### Test helpers

  def helper_confirm_sheet_contains_filename(self, sheet, filename):
    for row in sheet.rows:
      if row[0].value == filename:
        return True
    return False

  ##### Tests below here

  def test_can_create_simple_report_without_summary(self):
    # Edith imports a very short SPDX file as a new scan in the frotz
    # subproject frotz-dim
    result = runcmd(self, slm.cli, "frotz", "--subproject", "frotz-dim",
      "import-scan", PATH_SIMPLE_SPDX, "--scan_date", "2017-05-05",
      "--desc", "frotz-dim initial scan")
    self.assertEqual(0, result.exit_code)

    # Now Edith wants to get an xlsx report of the findings, sorted by
    # license category. She specifies the target output path and doesn't want
    # an initial summary sheet
    reportPath = self.reportDir.path + "/report.xlsx"
    result = runcmd(self, slm.cli, "frotz", "--subproject", "frotz-dim",
      "create-report", "--scan_id", "3", "--report_format", "xlsx",
      "--report_path", reportPath, "--no_summary")

    # The output message tells her it succeeded
    self.assertEqual(0, result.exit_code)
    self.assertEqual(f"Report successfully created at {reportPath}.\n", result.output)

    # She confirms that the file was created successfully
    self.assertTrue(os.path.isfile(reportPath))

    # Looking inside the workbook, she sees that the expected license
    # sheets are present
    wb = load_workbook(filename=reportPath)
    self.assertEqual(['Attribution', 'No license found'], wb.sheetnames)

    # and that the appropriate headers are present with the right fonts
    for sheet in wb:
      headerFile = sheet['A1']
      self.assertEqual("File", headerFile.value)
      self.assertEqual(16, headerFile.font.size)
      self.assertTrue(headerFile.font.bold)
      self.assertFalse(headerFile.alignment.wrap_text)
      headerLicense = sheet['B1']
      self.assertEqual("License", headerLicense.value)
      self.assertEqual(16, headerLicense.font.size)
      self.assertTrue(headerLicense.font.bold)
      self.assertFalse(headerLicense.alignment.wrap_text)

    # and that the file and license results are in the expected locations
    ws1 = wb["Attribution"]
    self.assertEqual("simple/file2.txt", ws1['A2'].value)
    self.assertEqual("MIT", ws1['B2'].value)
    ws2 = wb["No license found"]
    self.assertEqual("simple/dir1/subfile.txt", ws2['A2'].value)
    self.assertEqual("No license found", ws2['B2'].value)
    self.assertEqual("simple/file1.txt", ws2['A3'].value)
    self.assertEqual("No license found", ws2['B3'].value)
    self.assertEqual("simple/file3.txt", ws2['A4'].value)
    self.assertEqual("No license found", ws2['B4'].value)

    # and that the column widths are as expected
    self.assertEqual(100, ws2.column_dimensions["A"].width)
    self.assertEqual(60, ws2.column_dimensions["B"].width)

    # and that the file and license cells have the right fonts
    cellFile = ws2['A4']
    self.assertEqual(14, cellFile.font.size)
    self.assertFalse(cellFile.font.bold)
    self.assertTrue(cellFile.alignment.wrap_text)
    cellLicense = ws2['B4']
    self.assertEqual(14, cellLicense.font.size)
    self.assertFalse(cellLicense.font.bold)
    self.assertTrue(cellLicense.alignment.wrap_text)

  def test_can_create_report_with_summary(self):
    # Edith uses the existing frotz-nuclear scan in the sandbox. She requests
    # a license report, this time with a summary sheet (b/c she omits the
    # --no-summary flag)
    reportPath = self.reportDir.path + "/report.xlsx"
    result = runcmd(self, slm.cli, "frotz", "--subproject", "frotz-nuclear",
      "create-report", "--scan_id", "1", "--report_format", "xlsx",
      "--report_path", reportPath)

    # The output message tells her it succeeded
    self.assertEqual(0, result.exit_code)
    self.assertEqual(f"Report successfully created at {reportPath}.\n", result.output)

    # Looking inside the workbook, she sees that the first sheet is a
    # summary of all licenses and categories
    wb = load_workbook(filename=reportPath)
    self.assertEqual('License summary', wb.sheetnames[0])
    ws = wb['License summary']

    # and that the column widths are as expected
    self.assertEqual(3, ws.column_dimensions["A"].width)
    self.assertEqual(60, ws.column_dimensions["B"].width)
    self.assertEqual(10, ws.column_dimensions["C"].width)

    # and that the appropriate headers and wrapping are present
    headerA = ws['A1']
    self.assertEqual("License", headerA.value)
    self.assertEqual(16, headerA.font.size)
    self.assertTrue(headerA.font.bold)
    self.assertFalse(headerA.alignment.wrap_text)
    headerC = ws['C1']
    self.assertEqual("# of files", headerC.value)
    self.assertEqual(16, headerC.font.size)
    self.assertTrue(headerC.font.bold)
    self.assertFalse(headerA.alignment.wrap_text)

    # and that the category headers are appropriate
    catPL = ws['A3']
    self.assertEqual("Project Licenses:", catPL.value)
    self.assertEqual(16, catPL.font.size)
    self.assertTrue(catPL.font.bold)
    self.assertFalse(catPL.alignment.wrap_text)

    # and that the license and file count cells are as expected
    licApache = ws['B4']
    self.assertEqual("Apache-2.0", licApache.value)
    self.assertEqual(14, licApache.font.size)
    self.assertFalse(licApache.font.bold)
    self.assertTrue(licApache.alignment.wrap_text)
    filesApache = ws['C4']
    self.assertEqual(49, filesApache.value)
    self.assertEqual(14, filesApache.font.size)
    self.assertFalse(filesApache.font.bold)
    self.assertFalse(filesApache.alignment.wrap_text)

    # and that the Total and file count cells are as expected
    catTotal = ws['A9']
    self.assertEqual("TOTAL", catTotal.value)
    self.assertEqual(16, catTotal.font.size)
    self.assertTrue(catTotal.font.bold)
    self.assertFalse(catTotal.alignment.wrap_text)
    filesTotal = ws['C9']
    self.assertEqual(54, filesTotal.value)
    self.assertEqual(16, filesTotal.font.size)
    self.assertTrue(filesTotal.font.bold)
    self.assertFalse(filesTotal.alignment.wrap_text)

  def test_cannot_overwrite_existing_report_file(self):
    # Edith creates a report and then accidentally tries to overwrite it
    # with a second call to save to the same path
    reportPath = self.reportDir.path + "/report.xlsx"
    result = runcmd(self, slm.cli, "frotz", "--subproject", "frotz-nuclear",
      "create-report", "--scan_id", "1", "--report_format", "xlsx",
      "--report_path", reportPath)
    result = runcmd(self, slm.cli, "frotz", "--subproject", "frotz-nuclear",
      "create-report", "--scan_id", "1", "--report_format", "xlsx",
      "--report_path", reportPath)

    # It fails and explains why
    self.assertEqual(1, result.exit_code)
    self.assertEqual(f"File already exists at {reportPath} (use -f to force overwrite)\n", result.output)

  def test_can_overwrite_existing_report_file_with_force_flag(self):
    # Edith creates a report and then intentionally overwrites it
    # with a second call to save to the same path
    reportPath = self.reportDir.path + "/report.xlsx"
    result = runcmd(self, slm.cli, "frotz", "--subproject", "frotz-nuclear",
      "create-report", "--scan_id", "1", "--report_format", "xlsx",
      "--report_path", reportPath)
    result = runcmd(self, slm.cli, "frotz", "--subproject", "frotz-nuclear",
      "create-report", "--scan_id", "1", "--report_format", "xlsx",
      "--report_path", reportPath, "-f")

    # The output message tells her it succeeded
    self.assertEqual(0, result.exit_code)
    self.assertEqual(f"Report successfully created at {reportPath}.\n", result.output)

  def test_can_configure_to_omit_common_path_prefixes(self):
    # Edith configures the project so that reports will exclude the
    # common path prefixes
    result = runcmd(self, slm.cli, "frotz", "set-config",
      "analyze-exclude-path-prefix", "yes")
    self.assertEqual(0, result.exit_code)

    # She then creates a report
    reportPath = self.reportDir.path + "/report.xlsx"
    result = runcmd(self, slm.cli, "frotz", "--subproject", "frotz-nuclear",
      "create-report", "--scan_id", "1", "--report_format", "xlsx",
      "--report_path", reportPath, "--no_summary")
    self.assertEqual(0, result.exit_code)

    # Looking inside the workbook, she sees that the common license prefix
    # no longer shows up in the file listings
    wb = load_workbook(filename=reportPath)
    ws1 = wb["Project Licenses"]
    self.assertEqual("/.gitignore", ws1['A2'].value)
    for row in range(2, ws1.max_row+1):
      self.assertNotIn("master", ws1[f'A{row}'].value)

  def test_can_configure_to_label_files_with_certain_extensions(self):
    # Edith configures the project so that files with no license found,
    # but with certain file extensions, will be labeled separately in the
    # report
    result = runcmd(self, slm.cli, "frotz", "set-config",
      "analyze-extensions", "yes")
    self.assertEqual(0, result.exit_code)
    result = runcmd(self, slm.cli, "frotz", "set-config",
      "analyze-extensions-list", "json;jpg;png")
    self.assertEqual(0, result.exit_code)

    # She then imports the SPDX file
    result = runcmd(self, slm.cli, "frotz", "--subproject", "frotz-dim",
      "import-scan", PATH_SIMPLE_EXTENSION_SPDX, "--scan_date", "2017-05-05",
      "--desc", "frotz-dim scan to exclude extensions")
    self.assertEqual(0, result.exit_code)

    # She then creates a report
    reportPath = self.reportDir.path + "/report.xlsx"
    result = runcmd(self, slm.cli, "frotz", "--subproject", "frotz-dim",
      "create-report", "--scan_id", "3", "--report_format", "xlsx",
      "--report_path", reportPath)
    self.assertEqual(0, result.exit_code)

    # Looking inside the workbook, she sees that the summary displays a
    # separate line for these files
    wb = load_workbook(filename=reportPath)
    ws1 = wb["License summary"]
    flag_found = False
    for row in range(1, ws1.max_row+1):
      if ws1[f'B{row}'].value == "No license found - excluded file extension":
        flag_found = True
    self.assertTrue(flag_found)

    # And she sees that they are labeled separately in the file listing sheet
    ws2 = wb["No license found"]
    self.assertEqual("simple/dir1/subfile.txt", ws2['A2'].value)
    self.assertEqual("No license found", ws2['B2'].value)
    self.assertEqual("simple/a1.json", ws2['A5'].value)
    self.assertEqual("No license found - excluded file extension", ws2['B5'].value)

  def test_can_configure_to_label_empty_files(self):
    # Edith configures the project so that files with no license found,
    # which are totally empty, will be labeled separately in the report
    result = runcmd(self, slm.cli, "frotz", "set-config",
      "analyze-emptyfile", "yes")
    self.assertEqual(0, result.exit_code)

    # She then imports the SPDX file
    result = runcmd(self, slm.cli, "frotz", "--subproject", "frotz-dim",
      "import-scan", PATH_SIMPLE_EMPTYFILE_SPDX, "--scan_date", "2017-05-05",
      "--desc", "frotz-dim scan to label empty files")
    self.assertEqual(0, result.exit_code)

    # She then creates a report
    reportPath = self.reportDir.path + "/report.xlsx"
    result = runcmd(self, slm.cli, "frotz", "--subproject", "frotz-dim",
      "create-report", "--scan_id", "3", "--report_format", "xlsx",
      "--report_path", reportPath)
    self.assertEqual(0, result.exit_code)

    # Looking inside the workbook, she sees that the summary displays a
    # separate line for these files
    wb = load_workbook(filename=reportPath)
    ws1 = wb["License summary"]
    flag_found = False
    for row in range(1, ws1.max_row+1):
      if ws1[f'B{row}'].value == "No license found - empty file":
        flag_found = True
    self.assertTrue(flag_found)

    # And she sees that they are labeled separately in the file listing sheet
    ws2 = wb["No license found"]
    self.assertEqual("simple/a1.txt", ws2['A2'].value)
    self.assertEqual("No license found", ws2['B2'].value)
    self.assertEqual("simple/__init__.py", ws2['A5'].value)
    self.assertEqual("No license found - empty file", ws2['B5'].value)

  def test_can_configure_to_label_files_in_third_party_directories(self):
    # Edith configures the project so that files with no license found,
    # but in certain third party directories, will be labeled separately in the
    # report
    result = runcmd(self, slm.cli, "frotz", "set-config",
      "analyze-thirdparty", "yes")
    self.assertEqual(0, result.exit_code)
    result = runcmd(self, slm.cli, "frotz", "set-config",
      "analyze-thirdparty-dirs", "vendor;thirdparty")
    self.assertEqual(0, result.exit_code)

    # She then imports the SPDX file
    result = runcmd(self, slm.cli, "frotz", "--subproject", "frotz-dim",
      "import-scan", PATH_SIMPLE_THIRDPARTY_SPDX, "--scan_date", "2017-05-05",
      "--desc", "frotz-dim scan to exclude thirdparty dirs")
    self.assertEqual(0, result.exit_code)

    # She then creates a report
    reportPath = self.reportDir.path + "/report.xlsx"
    result = runcmd(self, slm.cli, "frotz", "--subproject", "frotz-dim",
      "create-report", "--scan_id", "3", "--report_format", "xlsx",
      "--report_path", reportPath)
    self.assertEqual(0, result.exit_code)

    # Looking inside the workbook, she sees that the summary displays a
    # separate line for these files
    wb = load_workbook(filename=reportPath)
    ws1 = wb["License summary"]
    flag_found = False
    for row in range(1, ws1.max_row+1):
      if ws1[f'B{row}'].value == "No license found - third party directory":
        flag_found = True
    self.assertTrue(flag_found)

    # And she sees that they are labeled separately in the file listing sheet
    ws2 = wb["No license found"]
    self.assertEqual("simple/dir1/subfile.txt", ws2['A2'].value)
    self.assertEqual("No license found", ws2['B2'].value)
    self.assertEqual("simple/file1.txt", ws2['A3'].value)
    self.assertEqual("No license found", ws2['B3'].value)
    self.assertEqual("simple/thirdparty/src/a1.json", ws2['A4'].value)
    self.assertEqual("No license found - third party directory", ws2['B4'].value)
    self.assertEqual("simple/vendor/file3.txt", ws2['A5'].value)
    self.assertEqual("No license found - third party directory", ws2['B5'].value)

  def test_can_configure_to_strip_licenseref_prefixes_in_xlsx_report(self):
    # Edith configures the project so that licenses beginning with "LicenseRef-"
    # will have that prefix stripped from the xlsx report
    result = runcmd(self, slm.cli, "frotz", "set-config",
      "report-strip-licenseref", "yes")
    self.assertEqual(0, result.exit_code)

    # She adds the LicenseRef- license
    result = runcmd(self, slm.cli, "frotz", "add-license",
      "LicenseRef-swinslow-1 AND Apache-2.0", "Other")
    self.assertEqual(0, result.exit_code)

    # She then imports the SPDX file
    result = runcmd(self, slm.cli, "frotz", "--subproject", "frotz-dim",
      "import-scan", PATH_SIMPLE_LICENSEREF_SPDX, "--scan_date", "2017-08-15",
      "--desc", "frotz-dim scan to strip LicenseRef- prefixes")
    self.assertEqual(0, result.exit_code)

    # She then creates a report
    reportPath = self.reportDir.path + "/report.xlsx"
    result = runcmd(self, slm.cli, "frotz", "--subproject", "frotz-dim",
      "create-report", "--scan_id", "3", "--report_format", "xlsx",
      "--report_path", reportPath)
    self.assertEqual(0, result.exit_code)

    # Looking inside the workbook, she sees that the summary displays
    # this license without the "LicenseRef-" prefix
    wb = load_workbook(filename=reportPath)
    ws1 = wb["License summary"]
    flag_found = False
    for row in range(1, ws1.max_row+1):
      if ws1[f'B{row}'].value == "swinslow-1 AND Apache-2.0":
        flag_found = True
    self.assertTrue(flag_found)

    # And she sees that it also has the prefix omitted in the file list sheet
    ws2 = wb["Other"]
    self.assertEqual("swinslow-1 AND Apache-2.0", ws2['B2'].value)

  def test_can_omit_report_path_and_get_default_location_and_name(self):
    # Edith chooses not to include a --report-path flag
    result = runcmd(self, slm.cli, "frotz",
      "create-report", "--scan_id", "1", "--report_format", "xlsx")

    # The output message tells her it succeeded, and that the report is in
    # the expected path and has the expected filename
    self.assertEqual(0, result.exit_code)
    expectedPath = os.path.join("projects", "frotz",
      "subprojects", "frotz-nuclear", "reports", "frotz-nuclear-2018-01-26.xlsx")
    fullPath = os.path.join(self.slmhome, expectedPath)
    self.assertEqual(f"Report successfully created at {fullPath}.\n", result.output)

    # She checks to make sure, and indeed the report is there
    checkForFileExists(self, self.slmhome, expectedPath)

  def test_cannot_omit_report_path_if_reporting_on_multiple_scans(self):
    # Edith chooses not to include a --report-path flag. However, she has included
    # multiple scans, and this causes things to fail because the appropriate
    # report name cannot be auto-determined.
    result = runcmd(self, slm.cli, "frotz",
      "create-report", "--scan_ids", "1,3", "--report_format", "xlsx")

    # It fails and explains why
    self.assertEqual(1, result.exit_code)
    self.assertEqual(f"Cannot auto-determine report path; --report_path must be included when multiple scans are included in one report.\n", result.output)

  def test_can_create_report_for_multiple_scans(self):
    # Edith imports a second SPDX file, for a different but related project,
    # as a new scan in the frotz subproject frotz-nuclear
    result = runcmd(self, slm.cli, "frotz", "add-license",
      "Apache-2.0 AND MIT", "Attribution")
    self.assertEqual(0, result.exit_code)
    result = runcmd(self, slm.cli, "frotz", "--subproject", "frotz-nuclear",
      "import-scan", PATH_SPDX_SUMMARIZER, "--scan_date", "2017-10-01",
      "--desc", "frotz-nuclear spdxSummarizer")
    self.assertEqual(0, result.exit_code)

    # Now Edith wants to get an xlsx report of the findings for BOTH scans,
    # combined into one report. She omits the summary
    reportPath = self.reportDir.path + "/report.xlsx"
    result = runcmd(self, slm.cli, "frotz", "--subproject", "frotz-nuclear",
      "create-report", "--scan_ids", "1,3", "--report_format", "xlsx",
      "--report_path", reportPath, "--no_summary")

    # The output message tells her it succeeded
    self.assertEqual(0, result.exit_code)
    self.assertEqual(f"Report successfully created at {reportPath}.\n", result.output)

    # She confirms that the file was created successfully
    self.assertTrue(os.path.isfile(reportPath))

    # Looking inside the workbook, she sees that the expected license
    # sheets are present
    wb = load_workbook(filename=reportPath)
    self.assertEqual(['Project Licenses', 'Attribution', 'No license found'], wb.sheetnames)

    # and that file results from both scans are found
    ws1 = wb["Project Licenses"]
    scan_1 = self.helper_confirm_sheet_contains_filename(ws1, "spdxLicenseManager-master/tests/ft_init.py")
    scan_3 = self.helper_confirm_sheet_contains_filename(ws1, "spdxSummarizer-master/LICENSE-docs.txt")
    self.assertTrue(scan_1)
    self.assertTrue(scan_3)

  # def test_can_create_report_for_multiple_scans_with_renamed_paths(self):
  #   # Same as prior test, but this time Edith wants to add a prefix before
  #   # all paths in scan 1
  #   result = runcmd(self, slm.cli, "frotz", "--subproject", "frotz-nuclear",
  #     "import-scan", PATH_SPDX_SUMMARIZER, "--scan_date", "2017-10-01",
  #     "--desc", "frotz-nuclear spdxSummarizer")
  #   self.assertEqual(0, result.exit_code)

  #   # Now she asks for the report, this time with the prefix noted
  #   reportPath = self.reportDir.path + "/report.xlsx"
  #   result = runcmd(self, slm.cli, "frotz", "--subproject", "frotz-nuclear",
  #     "create-report", "--scan_ids", "1:PREFIX,3", "--report_format", "xlsx",
  #     "--report_path", reportPath, "--no_summary")

  #   # The output message tells her it succeeded
  #   self.assertEqual(0, result.exit_code)
  #   self.assertEqual(f"Report successfully created at {reportPath}.\n", result.output)

  #   # She confirms that the file was created successfully
  #   self.assertTrue(os.path.isfile(reportPath))

  #   # Looking inside the workbook, she sees that the expected license
  #   # sheets are present
  #   wb = load_workbook(filename=reportPath)
  #   self.assertEqual(['Project Licenses', 'Attribution', 'No license found'], wb.sheetnames)

  #   # and that file results from both scans are found, with PREFIX/ added
  #   # to scan 1's files only
  #   ws1 = wb["Project Licenses"]
  #   scan_1_prefix = helper_confirm_sheet_contains_filename(ws1, "PREFIX/spdxLicenseManager-master/tests/ft_init.py")
  #   scan_1_noprefix = helper_confirm_sheet_contains_filename(ws1, "spdxLicenseManager-master/tests/ft_init.py")
  #   scan_3_prefix = helper_confirm_sheet_contains_filename(ws1, "PREFIX/spdxSummarizer-master/LICENSE-docs.txt")
  #   scan_3_noprefix = helper_confirm_sheet_contains_filename(ws1, "spdxLicenseManager-master/tests/ft_init.py")
  #   self.assertTrue(scan_1_prefix)
  #   self.assertFalse(scan_1_noprefix)
  #   self.assertFalse(scan_3_prefix)
  #   self.assertTrue(scan_3_noprefix)
