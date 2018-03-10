# tests/unit_reportexcel.py
#
# Unit test for spdxLicenseManager: creating Excel reports.
#
# Copyright (C) The Linux Foundation
# SPDX-License-Identifier: Apache-2.0
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import os
import unittest
from unittest import mock
from collections import OrderedDict
import datetime

from openpyxl import Workbook

from slm.datatypes import Category, File, License, Scan, Subproject
from slm.projectdb import ProjectDB, ProjectDBQueryError
from slm.reports.common import ReportFileError, ReportNotReadyError
from slm.reports.xlsx import ExcelReporter

class ReportExcelTestSuite(unittest.TestCase):
  """spdxLicenseManager Excel reporting unit test suite."""

  def setUp(self):
    # create and initialize an in-memory database
    self.db = ProjectDB()
    self.db.createDB(":memory:")
    self.db.initializeDBTables()

   # create reporter
    self.reporter = ExcelReporter(db=self.db)

  def tearDown(self):
    pass

  ##### Test helpers to mimic results from analysis

  def _buildCategories(self, results):
    self.cat2 = Category()
    self.cat2._id = 2
    self.cat2.name = "catID2"
    self.cat2.order = 1
    self.cat2.licensesSorted = OrderedDict()
    self.cat2.hasFiles = True
    results[2] = self.cat2

    # no files, cat3 should not appear in reports
    self.cat3 = Category()
    self.cat3._id = 3
    self.cat3.name = "catID3"
    self.cat3.order = 2
    self.cat3.licensesSorted = OrderedDict()
    self.cat3.hasFiles = False
    results[3] = self.cat3

    self.cat1 = Category()
    self.cat1._id = 1
    self.cat1.name = "catID1"
    self.cat1.order = 3
    self.cat1.licensesSorted = OrderedDict()
    self.cat1.hasFiles = True
    results[1] = self.cat1

    self.cat4 = Category()
    self.cat4._id = 4
    self.cat4.name = "No license found"
    self.cat4.order = 4
    self.cat4.licensesSorted = OrderedDict()
    self.cat4.hasFiles = True
    results[4] = self.cat4

  def _buildLicenses(self, results):
    # should sort alphabetically before lic1cat2 in reports
    self.lic2cat2 = License()
    self.lic2cat2._id = 2
    self.lic2cat2.name = "another lic2cat2"
    self.lic2cat2.category_id = 2
    self.lic2cat2.filesSorted = OrderedDict()
    self.lic2cat2.hasFiles = True
    self.cat2.licensesSorted[2] = self.lic2cat2

    self.lic1cat2 = License()
    self.lic1cat2._id = 1
    self.lic1cat2.name = "lic1cat2"
    self.lic1cat2.category_id = 2
    self.lic1cat2.filesSorted = OrderedDict()
    self.lic1cat2.hasFiles = True
    self.cat2.licensesSorted[1] = self.lic1cat2

    # no files, lic3cat2 should not appear in reports
    self.lic3cat2 = License()
    self.lic3cat2._id = 3
    self.lic3cat2.name = "nope lic3cat2"
    self.lic3cat2.category_id = 2
    self.lic3cat2.filesSorted = OrderedDict()
    self.lic3cat2.hasFiles = False
    self.cat2.licensesSorted[3] = self.lic3cat2

    # no files, lic4cat3 should not appear in reports
    self.lic4cat3 = License()
    self.lic4cat3._id = 4
    self.lic4cat3.name = "nope lic4cat3"
    self.lic4cat3.category_id = 3
    self.lic4cat3.filesSorted = OrderedDict()
    self.lic4cat3.hasFiles = False
    self.cat3.licensesSorted[4] = self.lic4cat3

    self.lic5cat1 = License()
    self.lic5cat1._id = 5
    self.lic5cat1.name = "a license"
    self.lic5cat1.category_id = 1
    self.lic5cat1.filesSorted = OrderedDict()
    self.lic5cat1.hasFiles = True
    self.cat1.licensesSorted[5] = self.lic5cat1

    # no license found option, for analysis outputs
    self.noLicFound = License()
    self.noLicFound._id = 6
    self.noLicFound.name = "No license found"
    self.noLicFound.category_id = 4
    self.noLicFound.filesSorted = OrderedDict()
    self.noLicFound.hasFiles = True
    self.cat4.licensesSorted[6] = self.noLicFound

  def _buildFiles(self, results):
    # should sort alphabetically by path before f1
    self.f2 = File()
    self.f2._id = 2
    self.f2.path = "/first/tmp/f2"
    self.f2.scan_id = 1
    self.f2.license_id = 2
    self.f2.findings = {}
    self.lic2cat2.filesSorted[2] = self.f2

    self.f1 = File()
    self.f1._id = 1
    self.f1.path = "/tmp/f1"
    self.f1.scan_id = 1
    self.f1.license_id = 2
    self.f1.findings = {}
    self.lic2cat2.filesSorted[1] = self.f1

    # should sort after f2 and f1 b/c license ID sorts after theirs
    self.f3 = File()
    self.f3._id = 3
    self.f3.path = "/earliest/tmp/f3"
    self.f3.scan_id = 1
    self.f3.license_id = 1
    self.f3.findings = {}
    self.lic1cat2.filesSorted[3] = self.f3

    self.f4 = File()
    self.f4._id = 4
    self.f4.path = "/tmp/f4"
    self.f4.scan_id = 1
    self.f4.license_id = 1
    self.f4.findings = {}
    self.lic1cat2.filesSorted[4] = self.f4

    self.f5 = File()
    self.f5._id = 5
    self.f5.path = "/tmp/f5"
    self.f5.scan_id = 1
    self.f5.license_id = 5
    self.f5.findings = {}
    self.lic5cat1.filesSorted[5] = self.f5

    self.f6 = File()
    self.f6._id = 6
    self.f6.path = "/tmp/f6.png"
    self.f6.scan_id = 1
    self.f6.license_id = 6
    self.f6.findings = {
      "extension": "yes",
    }
    self.noLicFound.filesSorted[6] = self.f6

    # add some more files for testing findings
    self.f7 = File()
    self.f7._id = 7
    self.f7.path = "/tmp/__init__.py"
    self.f7.scan_id = 1
    self.f7.license_id = 6
    self.f7.findings = {
      "emptyfile": "yes",
    }
    self.noLicFound.filesSorted[7] = self.f7

    self.f8 = File()
    self.f8._id = 8
    self.f8.path = "/tmp/vendor/dep.py"
    self.f8.scan_id = 1
    self.f8.license_id = 6
    self.f8.findings = {
      "thirdparty": "yes",
    }
    self.noLicFound.filesSorted[8] = self.f8

    self.f9 = File()
    self.f9._id = 9
    self.f9.path = "/tmp/code.py"
    self.f9.scan_id = 1
    self.f9.license_id = 6
    self.f9.findings = {}
    self.noLicFound.filesSorted[9] = self.f9

  def _getAnalysisResults(self):
    results = OrderedDict()
    self._buildCategories(results)
    self._buildLicenses(results)
    self._buildFiles(results)
    return results

  ##### Test cases below

  def test_new_reporter_is_in_reset_state(self):
    self.assertIsNone(self.reporter.results)
    self.assertIsNone(self.reporter.wb)
    self.assertFalse(self.reporter.reportGenerated)
    self.assertEqual({}, self.reporter.kwConfig)

  ##### Reporter config function tests

  def test_reporter_can_take_optional_config_params(self):
    configDict = {
      "include-summary": "yes",
    }
    newReporter = ExcelReporter(db=self.db, config=configDict)
    self.assertEqual("yes", newReporter.kwConfig["include-summary"])

  def test_can_get_reporter_final_config(self):
    self.db.setConfigValue(key="include-summary", value="yes")
    summary = self.reporter._getFinalConfigValue(key="include-summary")
    self.assertEqual(summary, "yes")

  def test_can_override_db_config_in_reporter_final_config(self):
    self.db.setConfigValue(key="include-summary", value="yes")
    newReporter = ExcelReporter(db=self.db, config={"include-summary": "no"})
    summary = newReporter._getFinalConfigValue(key="include-summary")
    self.assertEqual(summary, "no")

  ##### Reporter setup function tests

  def test_can_set_results_and_create_workbook(self):
    results = self._getAnalysisResults()
    self.reporter.setResults(results)
    self.assertEqual(Workbook, type(self.reporter.wb))
    self.assertEqual(results, self.reporter.results)

  ##### Reporter generate function tests

  def test_generate_report_fails_if_results_not_ready(self):
    with self.assertRaises(ReportNotReadyError):
      self.reporter.generate()
    self.assertFalse(self.reporter.reportGenerated)

    # even with workbook created, should fail if results aren't ready yet
    self.reporter.wb = Workbook()
    with self.assertRaises(ReportNotReadyError):
      self.reporter.generate()
    self.assertFalse(self.reporter.reportGenerated)

  @mock.patch('slm.reports.xlsx.ExcelReporter._generateFileListings')
  @mock.patch('slm.reports.xlsx.ExcelReporter._generateCategorySheets')
  def test_generate_calls_required_helpers(self, cs_mock, file_mock):
    results = self._getAnalysisResults()
    self.reporter.setResults(results)
    self.reporter.generate()
    cs_mock.assert_called_with(self.reporter.wb, self.reporter.results)
    file_mock.assert_called_with(self.reporter.wb, self.reporter.results)

  def test_generate_sets_report_generated_flag(self):
    results = self._getAnalysisResults()
    self.reporter.setResults(results)
    self.reporter.generate()
    self.assertTrue(self.reporter.reportGenerated)

  def test_can_generate_report_category_pages(self):
    wb = Workbook()
    results = self._getAnalysisResults()
    self.reporter._generateCategorySheets(wb, results)

    # workbook now has the expected category sheets
    # catID3 does not appear because it has no files
    self.assertEqual(["catID2", "catID1", "No license found"], wb.sheetnames)
    # proper headers are set for each category page
    # and header fonts are as expected
    for sheet in wb:
      headerFile = sheet['A1']
      self.assertEqual("File", headerFile.value)
      self.assertEqual(16, headerFile.font.size)
      self.assertTrue(headerFile.font.bold)
      self.assertFalse(headerFile.alignment.wrap_text)
      headerLicense = sheet['B1']
      self.assertEqual("License", sheet['B1'].value)
      self.assertEqual(16, headerLicense.font.size)
      self.assertTrue(headerLicense.font.bold)
      self.assertFalse(headerLicense.alignment.wrap_text)

      # column widths are as expected
      self.assertEqual(100, sheet.column_dimensions["A"].width)
      self.assertEqual(60, sheet.column_dimensions["B"].width)

  def test_cannot_generate_files_before_category_sheets_are_generated(self):
    wb = Workbook()
    results = self._getAnalysisResults()
    with self.assertRaises(ReportNotReadyError):
      self.reporter._generateFileListings(wb, results)

  def test_can_generate_file_listings(self):
    wb = Workbook()
    results = self._getAnalysisResults()
    self.reporter._generateCategorySheets(wb, results)
    self.reporter._generateFileListings(wb, results)

    # correct files and licenses are in the correct cells
    ws1 = wb['catID2']
    self.assertEqual("/first/tmp/f2", ws1['A2'].value)
    self.assertEqual("another lic2cat2", ws1['B2'].value)
    self.assertEqual("/tmp/f1", ws1['A3'].value)
    self.assertEqual("another lic2cat2", ws1['B3'].value)
    self.assertEqual("/earliest/tmp/f3", ws1['A4'].value)
    self.assertEqual("lic1cat2", ws1['B4'].value)
    self.assertEqual("/tmp/f4", ws1['A5'].value)
    self.assertEqual("lic1cat2", ws1['B5'].value)

    ws2 = wb['catID1']
    self.assertEqual("/tmp/f5", ws2['A2'].value)
    self.assertEqual("a license", ws2['B2'].value)

    cellFile = ws2['A2']
    self.assertEqual(14, cellFile.font.size)
    self.assertFalse(cellFile.font.bold)
    self.assertTrue(cellFile.alignment.wrap_text)
    cellLicense = ws2['B2']
    self.assertEqual(14, cellLicense.font.size)
    self.assertFalse(cellLicense.font.bold)
    self.assertTrue(cellLicense.alignment.wrap_text)

  @mock.patch('slm.reports.xlsx.ExcelReporter._generateSummarySheet')
  def test_summary_not_generated_if_config_not_set(self, summary_mock):
    results = self._getAnalysisResults()
    self.reporter.setResults(results)
    self.reporter.generate()
    summary_mock.assert_not_called()

  @mock.patch('slm.reports.xlsx.ExcelReporter._generateSummarySheet')
  def test_summary_is_generated_if_config_set(self, summary_mock):
    self.db.setConfigValue(key="include-summary", value="yes")
    results = self._getAnalysisResults()
    self.reporter.setResults(results)
    self.reporter.generate()
    summary_mock.assert_called_with(self.reporter.wb, self.reporter.results)

  def test_can_generate_summary_sheet(self):
    wb = Workbook()
    results = self._getAnalysisResults()
    self.reporter._generateSummarySheet(wb, results)
    # sheet title is as expected
    self.assertEqual('License summary', wb.sheetnames[0])
    ws = wb.active
    # column widths are as expected
    self.assertEqual(3, ws.column_dimensions["A"].width)
    self.assertEqual(60, ws.column_dimensions["B"].width)
    self.assertEqual(10, ws.column_dimensions["C"].width)
    # headers are as expected
    self.assertEqual("License", ws['A1'].value)
    self.assertEqual(16, ws['A1'].font.size)
    self.assertTrue(ws['A1'].font.bold)
    self.assertFalse(ws['A1'].alignment.wrap_text)
    self.assertEqual("# of files", ws['C1'].value)
    self.assertEqual(16, ws['C1'].font.size)
    self.assertTrue(ws['C1'].font.bold)
    self.assertFalse(ws['C1'].alignment.wrap_text)
    # and rows/values and formatting are all as expected
    # cats and licenses with no files should not appear
    self.assertEqual("catID2:", ws['A3'].value)
    self.assertEqual(16, ws['A3'].font.size)
    self.assertTrue(ws['A3'].font.bold)
    self.assertFalse(ws['A3'].alignment.wrap_text)
    self.assertEqual("another lic2cat2", ws['B4'].value)
    self.assertEqual(14, ws['B4'].font.size)
    self.assertFalse(ws['B4'].font.bold)
    self.assertTrue(ws['B4'].alignment.wrap_text)
    self.assertEqual(2, ws['C4'].value)
    self.assertEqual("lic1cat2", ws['B5'].value)
    self.assertEqual(2, ws['C5'].value)
    self.assertEqual("catID1:", ws['A6'].value)
    self.assertEqual("a license", ws['B7'].value)
    self.assertEqual(1, ws['C7'].value)
    self.assertEqual("No license found:", ws['A8'].value)
    self.assertEqual("No license found", ws['B9'].value)
    self.assertEqual(4, ws['C9'].value)
    self.assertEqual("TOTAL", ws['A11'].value)
    self.assertEqual(16, ws['A11'].font.size)
    self.assertTrue(ws['A11'].font.bold)
    self.assertFalse(ws['A11'].alignment.wrap_text)
    self.assertEqual(9, ws['C11'].value)
    self.assertEqual(16, ws['C11'].font.size)
    self.assertTrue(ws['C11'].font.bold)
    self.assertFalse(ws['C11'].alignment.wrap_text)

  def test_summary_sheet_not_overwritten_by_other_sheets(self):
    self.db.setConfigValue(key="include-summary", value="yes")
    results = self._getAnalysisResults()
    self.reporter.setResults(results)
    self.reporter.generate()
    # workbook should have the summary sheet AND the expected category sheets
    self.assertEqual(
      ["License summary", "catID2", "catID1", "No license found"],
      self.reporter.wb.sheetnames
    )
    # the summary sheet has the expected summary headers
    self.assertEqual("License", self.reporter.wb['License summary']['A1'].value)
    self.assertEqual("# of files", self.reporter.wb['License summary']['C1'].value)
    # and the file listing sheets have the expected file listing headers
    self.assertEqual("File", self.reporter.wb['catID2']['A1'].value)
    self.assertEqual("License", self.reporter.wb['catID2']['B1'].value)

  ##### Reporter license names for findings

  def test_can_modify_licenses_in_no_lic_found_cat_for_findings(self):
    self.db.setConfigValue(key="analyze-extensions", value="yes")
    results = self._getAnalysisResults()

    # hand the "No license found" category to the annotate function
    self.reporter._annotateNoLicenseFound(
      catNoLicense=self.cat4,
      nextLicID=7,
    )

    # check that it has been modified as expected, with new licenses,
    # and with the expected files within those licenses

    # this will temporarily get assigned the next available license ID
    # (only in memory, not committed to db)
    noLicExt = self.cat4.licensesSorted[7]
    self.assertEqual("No license found - excluded file extension", noLicExt.name)
    self.assertEqual(4, noLicExt.category_id)
    self.assertTrue(noLicExt.hasFiles)
    self.assertEqual(noLicExt.filesSorted[6], self.f6)

  ##### Reporter save function tests

  @mock.patch('slm.reports.xlsx.os.path.exists', return_value=True)
  def test_save_check_raises_exception_if_file_already_exists(self, os_exists):
    results = self._getAnalysisResults()
    self.reporter.setResults(results)
    self.reporter.generate()

    with self.assertRaises(ReportFileError):
      self.reporter._saveCheck(path="/tmp/fake/existing.xlsx")

  @mock.patch('slm.reports.xlsx.os.path.exists', return_value=True)
  @mock.patch('slm.reports.xlsx.os.access', return_value=True)
  def test_save_check_doesnt_raise_exception_if_replace_flag_set(self, os_access, os_exists):
    results = self._getAnalysisResults()
    self.reporter.setResults(results)
    self.reporter.generate()

    try:
      self.reporter._saveCheck(path="/tmp/fake/existing.xlsx", replace=True)
    except ReportFileError:
      self.fail("ReportFileError raised even though replace flag is True")

  @mock.patch('slm.reports.xlsx.os.access', return_value=False)
  def test_save_check_fails_if_no_write_permission_for_path(self, os_access):
    results = self._getAnalysisResults()
    self.reporter.setResults(results)
    self.reporter.generate()

    with self.assertRaises(ReportFileError):
      self.reporter._saveCheck(path="/tmp/readonly/report.xlsx")

  def test_save_check_fails_if_report_not_ready_yet(self):
    with self.assertRaises(ReportNotReadyError):
      self.reporter._saveCheck(path="whatever")

    # and even with results added, should fail if report not generated yet
    results = self._getAnalysisResults()
    self.reporter.setResults(results)
    with self.assertRaises(ReportNotReadyError):
      self.reporter._saveCheck(path="whatever")

  @mock.patch('slm.reports.xlsx.openpyxl.workbook.Workbook.save')
  @mock.patch('slm.reports.xlsx.ExcelReporter._saveCheck')
  def test_save_calls_save_checker(self, check_mock, save_mock):
    results = self._getAnalysisResults()
    self.reporter.setResults(results)
    self.reporter.generate()

    path = "something"
    replace = True
    self.reporter.save(path=path, replace=replace)
    check_mock.assert_called_with(path=path, replace=replace)
    save_mock.assert_called_with(path)

  ##### Reporter misc helper function tests

  def test_can_get_max_lic_id_from_results(self):
    results = self._getAnalysisResults()
    maxLicID = self.reporter._getResultsMaxLicenseID(results)
    self.assertEqual(6, maxLicID)
